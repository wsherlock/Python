import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('/Users/sherlock/Documents/python/test.xlsx')

print(wb.sheetnames)

sheet = wb['Sheet1']

print(sheet.max_column)
print(sheet.max_row)
print(sheet.title)

for row in sheet.rows:
    for cell in row:
        print(cell.value, end='\t')
    print('\n')